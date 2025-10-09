from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io
from typing import List

from app.database.connection import get_database
from app.models.financial import Receita, Despesa, Categoria


class ExcelService:
    """
    Serviço para exportação para Excel, formatado para máxima legibilidade
    e em conformidade com o PEP 8 (limite de 80 caracteres por linha).
    """

    # --- Estilos e Configurações Centralizados ---
    HEADER_FILL = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=12)
    TITLE_FONT = Font(bold=True, size=14, color="2F5496")
    TOTAL_FONT = Font(bold=True)
    TOTAL_FILL = PatternFill(
        start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
    )
    THIN_BORDER = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    HEADERS = {
        "Categorias": ['Nome', 'Tipo', 'Cor'],
        "Receitas": [
            'Data', 'Descrição', 'Categoria', 'Forma Recebimento', 'Valor'
        ],
        "Despesas": [
            'Data', 'Descrição', 'Categoria', 'Forma Pagamento', 'Valor'
        ]
    }
    
    COLUMN_WIDTHS = {
        "Categorias": {'A': 25, 'B': 15, 'C': 15},
        "Receitas": {'A': 15, 'B': 30, 'C': 20, 'D': 20, 'E': 15},
        "Despesas": {'A': 15, 'B': 30, 'C': 20, 'D': 20, 'E': 15},
        "Tutorial": {'A': 80}
    }

    def __init__(self):
        self.db = get_database()

    def _setup_sheet(
        self, workbook: Workbook, title: str, sheet_title: str, headers: List[str]
    ) -> Worksheet:
        worksheet = workbook.create_sheet(sheet_title)
        
        worksheet['A1'] = title
        worksheet['A1'].font = self.TITLE_FONT
        worksheet.merge_cells(
            start_row=1, start_column=1,
            end_row=1, end_column=len(headers)
        )
        
        for col_idx, header_text in enumerate(headers, start=1):
            cell = worksheet.cell(row=3, column=col_idx)
            cell.value = header_text
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.border = self.THIN_BORDER
            cell.alignment = Alignment(horizontal='center')
        
        if sheet_title in self.COLUMN_WIDTHS:
            for col_letter, width in self.COLUMN_WIDTHS[sheet_title].items():
                worksheet.column_dimensions[col_letter].width = width

        return worksheet

    async def export_excel(self, user_id: str) -> StreamingResponse:
        receitas_data = await self.db.receitas.find(
            {"user_id": user_id}
        ).to_list(None)
        despesas_data = await self.db.despesas.find(
            {"user_id": user_id}
        ).to_list(None)
        categorias_data = await self.db.categorias.find(
            {"user_id": user_id}
        ).to_list(None)
        
        receitas = [Receita(**data) for data in receitas_data]
        despesas = [Despesa(**data) for data in despesas_data]
        categorias = [Categoria(**data) for data in categorias_data]

        workbook = Workbook()
        
        self._create_tutorial_sheet(workbook)
        self._create_categorias_sheet(workbook, categorias)
        self._create_receitas_sheet(workbook, receitas)
        self._create_despesas_sheet(workbook, despesas)
        
        self._setup_sheet(
            workbook, "📊 RESUMO MENSAL", "Resumo Mensal", 
            headers=['Mês', 'Ano', 'Receitas', 'Despesas', 'Saldo', '% Economia']
        )
        self._setup_sheet(
            workbook, "🔮 PROJEÇÕES FINANCEIRAS", "Projeções",
            headers=['Mês', 'Receita Estimada', 'Despesa Estimada', 'Saldo Estimado']
        )
        self._setup_sheet(
            workbook, "📈 PAINEL DE CONTROLE", "Painel",
            headers=['Indicador', 'Valor', 'Meta', 'Status']
        )
        
        excel_file = io.BytesIO()
        workbook.save(excel_file)
        excel_file.seek(0)
        
        media_type = (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
        return StreamingResponse(
            excel_file,
            media_type=media_type,
            headers={
                "Content-Disposition": (
                    "attachment; filename=controle_financeiro.xlsx"
                )
            }
        )

    def _create_tutorial_sheet(self, workbook: Workbook):
        ws_tutorial = workbook.active
        ws_tutorial.title = "Tutorial"
        ws_tutorial.column_dimensions['A'].width = self.COLUMN_WIDTHS["Tutorial"]['A']
        
        ws_tutorial['A1'] = (
            "📚 BEM-VINDO À PLANILHA DE CONTROLE FINANCEIRO"
        )
        ws_tutorial['A1'].font = Font(bold=True, size=16, color="2F5496")
        
        instrucoes = [
            "", 
            "🎯 COMO USAR ESTA PLANILHA:", 
            "",
            "1️⃣ ABA 'RECEITAS': Preencha suas receitas mensais",
            "   - Células em BRANCO são editáveis",
            "   - Células em CINZA são calculadas automaticamente (NÃO EDITE)",
            "",
            # ... (restante das instruções)
        ]
        for i, texto in enumerate(instrucoes, start=2):
            ws_tutorial.cell(row=i, column=1, value=texto)
            if "️⃣" in texto or "⚠️" in texto:
                ws_tutorial.cell(row=i, column=1).font = Font(bold=True, size=11)
    
    def _create_categorias_sheet(
        self, workbook: Workbook, categorias: List[Categoria]
    ):
        headers = self.HEADERS["Categorias"]
        worksheet = self._setup_sheet(workbook, "📁 CATEGORIAS", "Categorias", headers)
        
        for row_idx, categoria in enumerate(categorias, start=4):
            worksheet.cell(row=row_idx, column=1, value=categoria.nome)
            worksheet.cell(row=row_idx, column=2, value=categoria.tipo.value)
            worksheet.cell(row=row_idx, column=3, value=categoria.cor)
            
            for col_idx in range(1, len(headers) + 1):
                worksheet.cell(row=row_idx, column=col_idx).border = self.THIN_BORDER

    def _create_receitas_sheet(
        self, workbook: Workbook, receitas: List[Receita]
    ):
        headers = self.HEADERS["Receitas"]
        worksheet = self._setup_sheet(workbook, "💰 RECEITAS", "Receitas", headers)
        
        for row_idx, receita in enumerate(receitas, start=4):
            worksheet.cell(row=row_idx, column=1, value=receita.data)
            worksheet.cell(row=row_idx, column=2, value=receita.descricao)
            worksheet.cell(row=row_idx, column=3, value=receita.categoria)
            worksheet.cell(row=row_idx, column=4, value=receita.forma_recebimento.value)
            worksheet.cell(row=row_idx, column=5, value=receita.valor)
            
            for col_idx in range(1, len(headers) + 1):
                worksheet.cell(row=row_idx, column=col_idx).border = self.THIN_BORDER
        
        ultima_linha = len(receitas) + 4
        worksheet.cell(row=ultima_linha, column=4, value="TOTAL:").font = self.TOTAL_FONT
        total_cell = worksheet.cell(row=ultima_linha, column=5)
        total_cell.value = f"=SUM(E4:E{ultima_linha-1})"
        total_cell.font = self.TOTAL_FONT
        total_cell.fill = self.TOTAL_FILL

    def _create_despesas_sheet(
        self, workbook: Workbook, despesas: List[Despesa]
    ):
        headers = self.HEADERS["Despesas"]
        worksheet = self._setup_sheet(workbook, "💸 DESPESAS", "Despesas", headers)
        
        for row_idx, despesa in enumerate(despesas, start=4):
            worksheet.cell(row=row_idx, column=1, value=despesa.data)
            worksheet.cell(row=row_idx, column=2, value=despesa.descricao)
            worksheet.cell(row=row_idx, column=3, value=despesa.categoria)
            worksheet.cell(row=row_idx, column=4, value=despesa.forma_pagamento.value)
            worksheet.cell(row=row_idx, column=5, value=despesa.valor)
            
            for col_idx in range(1, len(headers) + 1):
                worksheet.cell(row=row_idx, column=col_idx).border = self.THIN_BORDER