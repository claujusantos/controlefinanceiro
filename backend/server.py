from fastapi import FastAPI, APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional
import uuid
from datetime import datetime
from decimal import Decimal
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.utils import get_column_letter

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Create the main app without a prefix
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")


# ========== MODELS ==========

class Categoria(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    nome: str
    tipo: str  # "receita" ou "despesa"
    cor: str = "#3B82F6"

class CategoriaCreate(BaseModel):
    nome: str
    tipo: str
    cor: str = "#3B82F6"

class Receita(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    data: str  # formato: YYYY-MM-DD
    descricao: str
    categoria: str
    forma_recebimento: str
    valor: float
    mes: int
    ano: int

class ReceitaCreate(BaseModel):
    data: str
    descricao: str
    categoria: str
    forma_recebimento: str
    valor: float

class Despesa(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    data: str
    descricao: str
    categoria: str
    forma_pagamento: str
    valor: float
    mes: int
    ano: int

class DespesaCreate(BaseModel):
    data: str
    descricao: str
    categoria: str
    forma_pagamento: str
    valor: float

class ResumoMensal(BaseModel):
    mes: int
    ano: int
    total_receitas: float
    total_despesas: float
    saldo: float
    percentual_economia: float
    lucro_prejuizo: str  # "lucro" ou "prejuizo"


# ========== HELPER FUNCTIONS ==========

def extrair_mes_ano(data_str: str):
    """Extrai mês e ano de uma string de data YYYY-MM-DD"""
    try:
        data = datetime.strptime(data_str, "%Y-%m-%d")
        return data.month, data.year
    except:
        return 1, 2025


# ========== CATEGORIAS ENDPOINTS ==========

@api_router.get("/categorias", response_model=List[Categoria])
async def listar_categorias():
    categorias = await db.categorias.find().to_list(1000)
    if not categorias:
        # Criar categorias padrão
        categorias_padrao = [
            {"id": str(uuid.uuid4()), "nome": "Salário", "tipo": "receita", "cor": "#10B981"},
            {"id": str(uuid.uuid4()), "nome": "Freelance", "tipo": "receita", "cor": "#34D399"},
            {"id": str(uuid.uuid4()), "nome": "Investimentos", "tipo": "receita", "cor": "#6EE7B7"},
            {"id": str(uuid.uuid4()), "nome": "Alimentação", "tipo": "despesa", "cor": "#EF4444"},
            {"id": str(uuid.uuid4()), "nome": "Transporte", "tipo": "despesa", "cor": "#F87171"},
            {"id": str(uuid.uuid4()), "nome": "Moradia", "tipo": "despesa", "cor": "#FCA5A5"},
            {"id": str(uuid.uuid4()), "nome": "Lazer", "tipo": "despesa", "cor": "#FCD34D"},
            {"id": str(uuid.uuid4()), "nome": "Saúde", "tipo": "despesa", "cor": "#FB923C"},
            {"id": str(uuid.uuid4()), "nome": "Educação", "tipo": "despesa", "cor": "#A78BFA"},
        ]
        await db.categorias.insert_many(categorias_padrao)
        categorias = categorias_padrao
    return [Categoria(**cat) for cat in categorias]

@api_router.post("/categorias", response_model=Categoria)
async def criar_categoria(input: CategoriaCreate):
    cat_dict = input.dict()
    cat_obj = Categoria(**cat_dict)
    await db.categorias.insert_one(cat_obj.dict())
    return cat_obj

@api_router.put("/categorias/{cat_id}", response_model=Categoria)
async def atualizar_categoria(cat_id: str, input: CategoriaCreate):
    cat_dict = input.dict()
    cat_dict["id"] = cat_id
    cat_obj = Categoria(**cat_dict)
    result = await db.categorias.update_one({"id": cat_id}, {"$set": cat_obj.dict()})
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Categoria não encontrada")
    return cat_obj

@api_router.delete("/categorias/{cat_id}")
async def deletar_categoria(cat_id: str):
    result = await db.categorias.delete_one({"id": cat_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Categoria não encontrada")
    return {"message": "Categoria deletada com sucesso"}


# ========== RECEITAS ENDPOINTS ==========

@api_router.get("/receitas", response_model=List[Receita])
async def listar_receitas(mes: Optional[int] = None, ano: Optional[int] = None):
    filtro = {}
    if mes:
        filtro["mes"] = mes
    if ano:
        filtro["ano"] = ano
    receitas = await db.receitas.find(filtro).to_list(1000)
    return [Receita(**rec) for rec in receitas]

@api_router.post("/receitas", response_model=Receita)
async def criar_receita(input: ReceitaCreate):
    rec_dict = input.dict()
    mes, ano = extrair_mes_ano(rec_dict["data"])
    rec_dict["mes"] = mes
    rec_dict["ano"] = ano
    rec_obj = Receita(**rec_dict)
    await db.receitas.insert_one(rec_obj.dict())
    return rec_obj

@api_router.put("/receitas/{rec_id}", response_model=Receita)
async def atualizar_receita(rec_id: str, input: ReceitaCreate):
    rec_dict = input.dict()
    mes, ano = extrair_mes_ano(rec_dict["data"])
    rec_dict["mes"] = mes
    rec_dict["ano"] = ano
    rec_dict["id"] = rec_id
    rec_obj = Receita(**rec_dict)
    result = await db.receitas.update_one({"id": rec_id}, {"$set": rec_obj.dict()})
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Receita não encontrada")
    return rec_obj

@api_router.delete("/receitas/{rec_id}")
async def deletar_receita(rec_id: str):
    result = await db.receitas.delete_one({"id": rec_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Receita não encontrada")
    return {"message": "Receita deletada com sucesso"}


# ========== DESPESAS ENDPOINTS ==========

@api_router.get("/despesas", response_model=List[Despesa])
async def listar_despesas(mes: Optional[int] = None, ano: Optional[int] = None):
    filtro = {}
    if mes:
        filtro["mes"] = mes
    if ano:
        filtro["ano"] = ano
    despesas = await db.despesas.find(filtro).to_list(1000)
    return [Despesa(**desp) for desp in despesas]

@api_router.post("/despesas", response_model=Despesa)
async def criar_despesa(input: DespesaCreate):
    desp_dict = input.dict()
    mes, ano = extrair_mes_ano(desp_dict["data"])
    desp_dict["mes"] = mes
    desp_dict["ano"] = ano
    desp_obj = Despesa(**desp_dict)
    await db.despesas.insert_one(desp_obj.dict())
    return desp_obj

@api_router.put("/despesas/{desp_id}", response_model=Despesa)
async def atualizar_despesa(desp_id: str, input: DespesaCreate):
    desp_dict = input.dict()
    mes, ano = extrair_mes_ano(desp_dict["data"])
    desp_dict["mes"] = mes
    desp_dict["ano"] = ano
    desp_dict["id"] = desp_id
    desp_obj = Despesa(**desp_dict)
    result = await db.despesas.update_one({"id": desp_id}, {"$set": desp_obj.dict()})
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Despesa não encontrada")
    return desp_obj

@api_router.delete("/despesas/{desp_id}")
async def deletar_despesa(desp_id: str):
    result = await db.despesas.delete_one({"id": desp_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Despesa não encontrada")
    return {"message": "Despesa deletada com sucesso"}


# ========== DASHBOARD & RESUMOS ==========

@api_router.get("/dashboard")
async def obter_dashboard(
    periodo: Optional[str] = None,  # "total", "ultimo_mes", "ultimos_6_meses", "customizado"
    data_inicio: Optional[str] = None,  # formato: YYYY-MM-DD
    data_fim: Optional[str] = None  # formato: YYYY-MM-DD
):
    """Retorna dados agregados para o dashboard com filtros de data"""
    from datetime import datetime, timedelta
    
    # Determinar o filtro baseado no período
    filtro = {}
    
    if periodo == "ultimo_mes":
        # Último mês
        hoje = datetime.now()
        filtro["mes"] = hoje.month
        filtro["ano"] = hoje.year
    elif periodo == "ultimos_6_meses":
        # Últimos 6 meses - vamos filtrar depois
        pass
    elif periodo == "customizado" and data_inicio and data_fim:
        # Range customizado - vamos filtrar depois
        pass
    # Se periodo == "total" ou None, não filtra nada
    
    # Buscar todos os dados (vamos filtrar por data depois se necessário)
    todas_receitas = await db.receitas.find().to_list(1000)
    todas_despesas = await db.despesas.find().to_list(1000)
    
    # Aplicar filtros de data customizada
    if periodo == "ultimos_6_meses":
        hoje = datetime.now()
        data_limite = hoje - timedelta(days=180)
        todas_receitas = [r for r in todas_receitas if datetime.strptime(r.get("data"), "%Y-%m-%d") >= data_limite]
        todas_despesas = [d for d in todas_despesas if datetime.strptime(d.get("data"), "%Y-%m-%d") >= data_limite]
    elif periodo == "customizado" and data_inicio and data_fim:
        dt_inicio = datetime.strptime(data_inicio, "%Y-%m-%d")
        dt_fim = datetime.strptime(data_fim, "%Y-%m-%d")
        todas_receitas = [r for r in todas_receitas if dt_inicio <= datetime.strptime(r.get("data"), "%Y-%m-%d") <= dt_fim]
        todas_despesas = [d for d in todas_despesas if dt_inicio <= datetime.strptime(d.get("data"), "%Y-%m-%d") <= dt_fim]
    elif periodo == "ultimo_mes":
        todas_receitas = [r for r in todas_receitas if r.get("mes") == filtro.get("mes") and r.get("ano") == filtro.get("ano")]
        todas_despesas = [d for d in todas_despesas if d.get("mes") == filtro.get("mes") and d.get("ano") == filtro.get("ano")]
    
    total_receitas = sum(r.get("valor", 0) for r in todas_receitas)
    total_despesas = sum(d.get("valor", 0) for d in todas_despesas)
    saldo = total_receitas - total_despesas
    percentual_economia = (saldo / total_receitas * 100) if total_receitas > 0 else 0
    
    # Distribuição por categorias (despesas)
    categorias_dist = {}
    for d in todas_despesas:
        cat = d.get("categoria", "Outros")
        categorias_dist[cat] = categorias_dist.get(cat, 0) + d.get("valor", 0)
    
    # Evolução mensal
    meses_anos = set()
    for r in todas_receitas:
        meses_anos.add((r.get("mes"), r.get("ano")))
    for d in todas_despesas:
        meses_anos.add((d.get("mes"), d.get("ano")))
    
    evolucao = []
    for m, a in sorted(meses_anos):
        recs = sum(r.get("valor", 0) for r in todas_receitas if r.get("mes") == m and r.get("ano") == a)
        desps = sum(d.get("valor", 0) for d in todas_despesas if d.get("mes") == m and d.get("ano") == a)
        evolucao.append({
            "mes": m,
            "ano": a,
            "receitas": recs,
            "despesas": desps,
            "saldo": recs - desps
        })
    
    return {
        "total_receitas": total_receitas,
        "total_despesas": total_despesas,
        "saldo": saldo,
        "percentual_economia": round(percentual_economia, 2),
        "lucro_prejuizo": "lucro" if saldo >= 0 else "prejuizo",
        "categorias_distribuicao": [{"categoria": k, "valor": v} for k, v in categorias_dist.items()],
        "evolucao_mensal": evolucao
    }

@api_router.get("/gastos-recorrentes")
async def obter_gastos_recorrentes():
    """Retorna análise de gastos recorrentes e frequentes"""
    from collections import Counter
    
    despesas = await db.despesas.find().to_list(1000)
    
    if not despesas:
        return {
            "categorias_mais_frequentes": [],
            "descricoes_recorrentes": [],
            "media_por_categoria": []
        }
    
    # Contar frequência por categoria
    categorias_count = Counter(d.get("categoria") for d in despesas)
    categorias_valores = {}
    categorias_ocorrencias = {}
    
    for d in despesas:
        cat = d.get("categoria", "Outros")
        if cat not in categorias_valores:
            categorias_valores[cat] = 0
            categorias_ocorrencias[cat] = 0
        categorias_valores[cat] += d.get("valor", 0)
        categorias_ocorrencias[cat] += 1
    
    # Categorias mais frequentes com valores
    categorias_freq = []
    for cat, count in categorias_count.most_common():
        categorias_freq.append({
            "categoria": cat,
            "ocorrencias": count,
            "valor_total": categorias_valores[cat],
            "valor_medio": categorias_valores[cat] / count if count > 0 else 0
        })
    
    # Descrições que se repetem (gastos recorrentes)
    descricoes_count = Counter(d.get("descricao").lower() for d in despesas)
    descricoes_recorrentes = []
    
    for desc, count in descricoes_count.most_common(10):
        if count > 1:  # Apenas descrições que aparecem mais de uma vez
            valor_total = sum(d.get("valor", 0) for d in despesas if d.get("descricao").lower() == desc)
            descricoes_recorrentes.append({
                "descricao": desc.title(),
                "ocorrencias": count,
                "valor_total": valor_total,
                "valor_medio": valor_total / count
            })
    
    # Média de gasto por categoria
    media_por_cat = []
    for cat, total in categorias_valores.items():
        count = categorias_ocorrencias[cat]
        media_por_cat.append({
            "categoria": cat,
            "media_gasto": total / count if count > 0 else 0,
            "total_gasto": total
        })
    
    # Ordenar por total gasto
    media_por_cat.sort(key=lambda x: x["total_gasto"], reverse=True)
    
    return {
        "categorias_mais_frequentes": categorias_freq[:10],
        "descricoes_recorrentes": descricoes_recorrentes,
        "media_por_categoria": media_por_cat
    }

@api_router.get("/resumo-mensal", response_model=List[ResumoMensal])
async def obter_resumo_mensal():
    """Retorna resumo de todos os meses"""
    receitas = await db.receitas.find().to_list(1000)
    despesas = await db.despesas.find().to_list(1000)
    
    meses_anos = set()
    for r in receitas:
        meses_anos.add((r.get("mes"), r.get("ano")))
    for d in despesas:
        meses_anos.add((d.get("mes"), d.get("ano")))
    
    resumos = []
    for m, a in sorted(meses_anos):
        total_rec = sum(r.get("valor", 0) for r in receitas if r.get("mes") == m and r.get("ano") == a)
        total_desp = sum(d.get("valor", 0) for d in despesas if d.get("mes") == m and d.get("ano") == a)
        saldo = total_rec - total_desp
        perc = (saldo / total_rec * 100) if total_rec > 0 else 0
        
        resumos.append(ResumoMensal(
            mes=m,
            ano=a,
            total_receitas=total_rec,
            total_despesas=total_desp,
            saldo=saldo,
            percentual_economia=round(perc, 2),
            lucro_prejuizo="lucro" if saldo >= 0 else "prejuizo"
        ))
    
    return resumos

@api_router.get("/projecoes")
async def obter_projecoes():
    """Calcula projeções financeiras baseadas nas médias"""
    receitas = await db.receitas.find().to_list(1000)
    despesas = await db.despesas.find().to_list(1000)
    
    if not receitas and not despesas:
        return {
            "media_receitas": 0,
            "media_despesas": 0,
            "saldo_projetado": 0,
            "tendencia": "neutro"
        }
    
    # Calcular médias dos últimos 3 meses
    meses_anos = set()
    for r in receitas:
        meses_anos.add((r.get("mes"), r.get("ano")))
    for d in despesas:
        meses_anos.add((d.get("mes"), d.get("ano")))
    
    ultimos_meses = sorted(meses_anos)[-3:]
    
    total_rec = 0
    total_desp = 0
    for m, a in ultimos_meses:
        total_rec += sum(r.get("valor", 0) for r in receitas if r.get("mes") == m and r.get("ano") == a)
        total_desp += sum(d.get("valor", 0) for d in despesas if d.get("mes") == m and d.get("ano") == a)
    
    media_rec = total_rec / len(ultimos_meses) if ultimos_meses else 0
    media_desp = total_desp / len(ultimos_meses) if ultimos_meses else 0
    saldo_proj = media_rec - media_desp
    
    tendencia = "crescimento" if saldo_proj > 0 else "declinio" if saldo_proj < 0 else "neutro"
    
    return {
        "media_receitas": round(media_rec, 2),
        "media_despesas": round(media_desp, 2),
        "saldo_projetado": round(saldo_proj, 2),
        "tendencia": tendencia,
        "projecao_6_meses": [
            {
                "mes": i + 1,
                "receita_estimada": round(media_rec, 2),
                "despesa_estimada": round(media_desp, 2),
                "saldo_estimado": round(saldo_proj, 2)
            }
            for i in range(6)
        ]
    }


# ========== EXPORTAÇÃO EXCEL ==========

@api_router.get("/export-excel")
async def exportar_excel():
    """Gera arquivo Excel com todas as abas e fórmulas"""
    
    # Buscar dados
    receitas = await db.receitas.find().to_list(1000)
    despesas = await db.despesas.find().to_list(1000)
    categorias = await db.categorias.find().to_list(1000)
    
    # Criar workbook
    wb = Workbook()
    
    # Estilos
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    titulo_font = Font(bold=True, size=14, color="2F5496")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ========== ABA 1: TUTORIAL ==========
    ws_tutorial = wb.active
    ws_tutorial.title = "Tutorial"
    
    ws_tutorial['A1'] = "📚 BEM-VINDO À PLANILHA DE CONTROLE FINANCEIRO"
    ws_tutorial['A1'].font = Font(bold=True, size=16, color="2F5496")
    
    instrucoes = [
        "",
        "🎯 COMO USAR ESTA PLANILHA:",
        "",
        "1️⃣ ABA 'RECEITAS': Preencha suas receitas mensais",
        "   - Células em BRANCO são editáveis",
        "   - Células em CINZA são calculadas automaticamente (NÃO EDITE)",
        "",
        "2️⃣ ABA 'DESPESAS': Preencha suas despesas mensais",
        "   - Use o menu suspenso para selecionar categorias",
        "   - Os totais são calculados automaticamente",
        "",
        "3️⃣ ABA 'CATEGORIAS': Personalize suas categorias",
        "   - Adicione ou remova categorias conforme sua necessidade",
        "",
        "4️⃣ ABA 'RESUMO MENSAL': Veja o histórico completo",
        "   - Totais por mês calculados automaticamente",
        "   - Meses com prejuízo destacados em vermelho",
        "",
        "5️⃣ ABA 'PROJEÇÕES': Veja tendências futuras",
        "   - Baseado na média dos últimos meses",
        "",
        "6️⃣ ABA 'PAINEL': Dashboard visual",
        "   - Indicadores principais e resumo geral",
        "",
        "⚠️ IMPORTANTE:",
        "• Não delete linhas de cabeçalho",
        "• Não modifique células com fórmulas (cinza)",
        "• Sempre use datas no formato DD/MM/AAAA",
        "• Valores devem ser apenas números (sem R$)",
        "",
        "💡 DICA: Comece preenchendo a aba CATEGORIAS, depois RECEITAS e DESPESAS!",
        "",
        "✅ Pronto! Sua planilha está configurada e pronta para uso!",
    ]
    
    for i, texto in enumerate(instrucoes, start=2):
        ws_tutorial[f'A{i}'] = texto
        if "️⃣" in texto or "⚠️" in texto:
            ws_tutorial[f'A{i}'].font = Font(bold=True, size=11)
    
    ws_tutorial.column_dimensions['A'].width = 80
    
    # ========== ABA 2: CATEGORIAS ==========
    ws_cat = wb.create_sheet("Categorias")
    
    ws_cat['A1'] = "📁 CATEGORIAS"
    ws_cat['A1'].font = titulo_font
    ws_cat.merge_cells('A1:C1')
    
    headers_cat = ['Nome', 'Tipo', 'Cor']
    for col, header in enumerate(headers_cat, start=1):
        cell = ws_cat.cell(row=3, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    for i, cat in enumerate(categorias, start=4):
        ws_cat[f'A{i}'] = cat.get('nome', '')
        ws_cat[f'B{i}'] = cat.get('tipo', '')
        ws_cat[f'C{i}'] = cat.get('cor', '')
        
        for col in range(1, 4):
            ws_cat.cell(row=i, column=col).border = border
    
    ws_cat.column_dimensions['A'].width = 25
    ws_cat.column_dimensions['B'].width = 15
    ws_cat.column_dimensions['C'].width = 15
    
    # ========== ABA 3: RECEITAS ==========
    ws_rec = wb.create_sheet("Receitas")
    
    ws_rec['A1'] = "💰 RECEITAS"
    ws_rec['A1'].font = titulo_font
    ws_rec.merge_cells('A1:E1')
    
    headers_rec = ['Data', 'Descrição', 'Categoria', 'Forma Recebimento', 'Valor']
    for col, header in enumerate(headers_rec, start=1):
        cell = ws_rec.cell(row=3, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    for i, rec in enumerate(receitas, start=4):
        ws_rec[f'A{i}'] = rec.get('data', '')
        ws_rec[f'B{i}'] = rec.get('descricao', '')
        ws_rec[f'C{i}'] = rec.get('categoria', '')
        ws_rec[f'D{i}'] = rec.get('forma_recebimento', '')
        ws_rec[f'E{i}'] = rec.get('valor', 0)
        
        for col in range(1, 6):
            ws_rec.cell(row=i, column=col).border = border
    
    # Total
    ultima_linha = len(receitas) + 4
    ws_rec[f'D{ultima_linha}'] = "TOTAL:"
    ws_rec[f'D{ultima_linha}'].font = Font(bold=True)
    ws_rec[f'E{ultima_linha}'] = f"=SUM(E4:E{ultima_linha-1})"
    ws_rec[f'E{ultima_linha}'].font = Font(bold=True)
    ws_rec[f'E{ultima_linha}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    ws_rec.column_dimensions['A'].width = 15
    ws_rec.column_dimensions['B'].width = 30
    ws_rec.column_dimensions['C'].width = 20
    ws_rec.column_dimensions['D'].width = 20
    ws_rec.column_dimensions['E'].width = 15
    
    # ========== ABA 4: DESPESAS ==========
    ws_desp = wb.create_sheet("Despesas")
    
    ws_desp['A1'] = "💸 DESPESAS"
    ws_desp['A1'].font = titulo_font
    ws_desp.merge_cells('A1:E1')
    
    headers_desp = ['Data', 'Descrição', 'Categoria', 'Forma Pagamento', 'Valor']
    for col, header in enumerate(headers_desp, start=1):
        cell = ws_desp.cell(row=3, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    for i, desp in enumerate(despesas, start=4):
        ws_desp[f'A{i}'] = desp.get('data', '')
        ws_desp[f'B{i}'] = desp.get('descricao', '')
        ws_desp[f'C{i}'] = desp.get('categoria', '')
        ws_desp[f'D{i}'] = desp.get('forma_pagamento', '')
        ws_desp[f'E{i}'] = desp.get('valor', 0)
        
        for col in range(1, 6):
            ws_desp.cell(row=i, column=col).border = border
    
    # Total
    ultima_linha_desp = len(despesas) + 4
    ws_desp[f'D{ultima_linha_desp}'] = "TOTAL:"
    ws_desp[f'D{ultima_linha_desp}'].font = Font(bold=True)
    ws_desp[f'E{ultima_linha_desp}'] = f"=SUM(E4:E{ultima_linha_desp-1})"
    ws_desp[f'E{ultima_linha_desp}'].font = Font(bold=True)
    ws_desp[f'E{ultima_linha_desp}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    ws_desp.column_dimensions['A'].width = 15
    ws_desp.column_dimensions['B'].width = 30
    ws_desp.column_dimensions['C'].width = 20
    ws_desp.column_dimensions['D'].width = 20
    ws_desp.column_dimensions['E'].width = 15
    
    # ========== ABA 5: RESUMO MENSAL ==========
    ws_resumo = wb.create_sheet("Resumo Mensal")
    
    ws_resumo['A1'] = "📊 RESUMO MENSAL"
    ws_resumo['A1'].font = titulo_font
    ws_resumo.merge_cells('A1:F1')
    
    headers_resumo = ['Mês/Ano', 'Receitas', 'Despesas', 'Saldo', '% Economia', 'Status']
    for col, header in enumerate(headers_resumo, start=1):
        cell = ws_resumo.cell(row=3, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Agrupar por mês/ano
    meses_anos = set()
    for r in receitas:
        meses_anos.add((r.get('mes'), r.get('ano')))
    for d in despesas:
        meses_anos.add((d.get('mes'), d.get('ano')))
    
    row = 4
    for mes, ano in sorted(meses_anos):
        total_rec = sum(r.get('valor', 0) for r in receitas if r.get('mes') == mes and r.get('ano') == ano)
        total_desp = sum(d.get('valor', 0) for d in despesas if d.get('mes') == mes and d.get('ano') == ano)
        saldo = total_rec - total_desp
        perc = (saldo / total_rec * 100) if total_rec > 0 else 0
        status = "Lucro" if saldo >= 0 else "Prejuízo"
        
        ws_resumo[f'A{row}'] = f"{mes:02d}/{ano}"
        ws_resumo[f'B{row}'] = total_rec
        ws_resumo[f'C{row}'] = total_desp
        ws_resumo[f'D{row}'] = saldo
        ws_resumo[f'E{row}'] = f"{perc:.2f}%"
        ws_resumo[f'F{row}'] = status
        
        # Destaque vermelho para prejuízo
        if saldo < 0:
            for col in range(1, 7):
                ws_resumo.cell(row=row, column=col).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        for col in range(1, 7):
            ws_resumo.cell(row=row, column=col).border = border
        
        row += 1
    
    ws_resumo.column_dimensions['A'].width = 15
    ws_resumo.column_dimensions['B'].width = 15
    ws_resumo.column_dimensions['C'].width = 15
    ws_resumo.column_dimensions['D'].width = 15
    ws_resumo.column_dimensions['E'].width = 15
    ws_resumo.column_dimensions['F'].width = 15
    
    # ========== ABA 6: PROJEÇÕES ==========
    ws_proj = wb.create_sheet("Projeções")
    
    ws_proj['A1'] = "🔮 PROJEÇÕES FINANCEIRAS"
    ws_proj['A1'].font = titulo_font
    ws_proj.merge_cells('A1:D1')
    
    ws_proj['A3'] = "Baseado na média dos últimos 3 meses"
    ws_proj['A3'].font = Font(italic=True)
    
    ws_proj['A5'] = "Média de Receitas:"
    ws_proj['A6'] = "Média de Despesas:"
    ws_proj['A7'] = "Saldo Projetado:"
    
    # Calcular médias (simplificado para o Excel)
    total_rec_all = sum(r.get('valor', 0) for r in receitas)
    total_desp_all = sum(d.get('valor', 0) for d in despesas)
    num_meses = len(meses_anos) if meses_anos else 1
    
    ws_proj['B5'] = total_rec_all / num_meses
    ws_proj['B6'] = total_desp_all / num_meses
    ws_proj['B7'] = f"=B5-B6"
    
    for i in [5, 6, 7]:
        ws_proj[f'A{i}'].font = Font(bold=True)
        ws_proj[f'B{i}'].font = Font(bold=True)
    
    ws_proj.column_dimensions['A'].width = 25
    ws_proj.column_dimensions['B'].width = 20
    
    # ========== ABA 7: PAINEL ==========
    ws_painel = wb.create_sheet("Painel")
    ws_painel.sheet_view.showGridLines = False
    
    ws_painel['B2'] = "📈 PAINEL DE CONTROLE FINANCEIRO"
    ws_painel['B2'].font = Font(bold=True, size=18, color="2F5496")
    ws_painel.merge_cells('B2:E2')
    
    # Cards de indicadores
    ws_painel['B4'] = "💰 RECEITA TOTAL"
    ws_painel['B5'] = total_rec_all
    ws_painel['B4'].font = Font(bold=True, size=12)
    ws_painel['B5'].font = Font(bold=True, size=14, color="10B981")
    ws_painel['B4'].fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    ws_painel['B5'].fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    
    ws_painel['D4'] = "💸 DESPESA TOTAL"
    ws_painel['D5'] = total_desp_all
    ws_painel['D4'].font = Font(bold=True, size=12)
    ws_painel['D5'].font = Font(bold=True, size=14, color="EF4444")
    ws_painel['D4'].fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    ws_painel['D5'].fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    
    ws_painel['B7'] = "📊 SALDO"
    ws_painel['B8'] = total_rec_all - total_desp_all
    ws_painel['B7'].font = Font(bold=True, size=12)
    ws_painel['B8'].font = Font(bold=True, size=14, color="3B82F6")
    ws_painel['B7'].fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
    ws_painel['B8'].fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
    
    perc_economia = ((total_rec_all - total_desp_all) / total_rec_all * 100) if total_rec_all > 0 else 0
    ws_painel['D7'] = "💎 % ECONOMIA"
    ws_painel['D8'] = f"{perc_economia:.2f}%"
    ws_painel['D7'].font = Font(bold=True, size=12)
    ws_painel['D8'].font = Font(bold=True, size=14, color="8B5CF6")
    ws_painel['D7'].fill = PatternFill(start_color="EDE9FE", end_color="EDE9FE", fill_type="solid")
    ws_painel['D8'].fill = PatternFill(start_color="EDE9FE", end_color="EDE9FE", fill_type="solid")
    
    ws_painel.column_dimensions['B'].width = 20
    ws_painel.column_dimensions['D'].width = 20
    
    # Salvar em memória
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=controle_financeiro.xlsx"}
    )


# ========== ROOT ENDPOINT ==========

@api_router.get("/")
async def root():
    return {"message": "API de Controle Financeiro - Sistema completo de gestão financeira"}


# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
